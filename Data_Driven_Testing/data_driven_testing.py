# Data Driver Testing basically means testing our test script with different sets of data drived from external file like .xlsx or .csv file.
# By default the selenium-webdriver doesn't support the interaction with .xslx or .csv files, I need to import related 3rd party library or package to interact.
# openpyxl is a package/library that is used to with selenium-webdriver to interact with the .xlsx files.

import openpyxl,os

# steps to read the data from the Excel file
# step1
file = os.getcwd()+"/data_file.xlsx"

# step2
workbook = openpyxl.load_workbook(file)

# step3
sheet = workbook["Sheet1"] # I need to provide whatever the sheet name is.
# Note, if there is only one sheet present in our Excel workbook, with whatever name, then I can use
# sheet = workbook.active

# step4
rows = sheet.max_row
print(rows)
columns = sheet.max_column
print(columns)

# step5
for r in range(2,rows+1):   # w.r.t excel the rows and column count starts from 1.
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end='    ')
    print()  # To get the cursor to next line